import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def extract_matrix_to_xlsx(xml_path: str, xlsx_out: str):
    tree = ET.parse(xml_path)
    root = tree.getroot()

    clash_tests = []
    all_rules = set()

    # --- Первый проход: собираем тесты и правила ---
    for clashtest in root.findall(".//clashtest"):
        clash_name = clashtest.get("name")
        if not clash_name:
            continue

        rules_block = clashtest.find("rules")

        enabled_rules = set()
        if rules_block is not None:
            for rule in rules_block.findall("rule"):
                rule_name = rule.get("name")
                if rule_name:
                    all_rules.add(rule_name)
                    enabled_rules.add(rule_name)

        clash_tests.append((clash_name, enabled_rules))

    # отсортируем правила
    all_rules = sorted(all_rules)

    # --- Создаём Excel-файл ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Clash Matrix"

    # первая строка: заголовки
    ws.cell(row=1, column=1, value="Clash Test")

    for col, rule in enumerate(all_rules, start=2):
        ws.cell(row=1, column=col, value=rule)

    # данные
    for row_index, (clash_name, enabled_rules) in enumerate(clash_tests, start=2):
        ws.cell(row=row_index, column=1, value=clash_name)

        for col_index, rule in enumerate(all_rules, start=2):
            ws.cell(
                row=row_index,
                column=col_index,
                value="✓" if rule in enabled_rules else ""
            )

    # автоширина колонок
    for col in range(1, len(all_rules) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 25

    wb.save(xlsx_out)
    print(f"Готово! Файл сохранён: {xlsx_out}")


if __name__ == "__main__":
    xml_file = r"D:\загрузки\!К32_SM_Block_B_GF_увязка (ПЧ)_v3.xml"
    out_xlsx = r"D:\загрузки\clash_matrix.xlsx"

    if not Path(xml_file).exists():
        print(f"Файл не найден: {xml_file}")
    else:
        extract_matrix_to_xlsx(xml_file, out_xlsx)
